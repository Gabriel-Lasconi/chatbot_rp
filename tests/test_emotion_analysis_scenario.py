import unittest
from openpyxl import Workbook
from app.chatbot_generative import ChatbotGenerative


class TestEmotionAnalysisToExcel(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """
        Create and initialize the ChatbotGenerative instance once for the entire test class.
        Also create an Excel workbook/sheet for storing results.
        """
        cls.chatbot = ChatbotGenerative()

        # Prepare an in-memory workbook
        cls.wb = Workbook()
        cls.sheet = cls.wb.active
        cls.sheet.title = "Emotion Analysis Results"

        # Write header row
        cls.sheet.cell(row=1, column=1, value="Scenario")
        cls.sheet.cell(row=1, column=2, value="Line Content")
        cls.sheet.cell(row=1, column=3, value="Expected Emotion")
        cls.sheet.cell(row=1, column=4, value="Detected Emotion")
        cls.sheet.cell(row=1, column=5, value="Correct?")

        cls.current_row = 2  # to track where to write next

        # This will count correct lines vs total
        cls.total_lines = 0
        cls.correct_count = 0

    def setUp(self):
        """
        Reset the chatbot conversation for each scenario.
        """
        self.chatbot.reset_conversation()

    @classmethod
    def tearDownClass(cls):
        """
        After all tests, compute accuracy, write to Excel, and save the file.
        """
        # Calculate accuracy
        if cls.total_lines > 0:
            accuracy = (cls.correct_count / cls.total_lines) * 100
        else:
            accuracy = 0.0

        # Optionally write accuracy to the sheet
        last_row = cls.current_row + 2
        cls.sheet.cell(row=last_row, column=1, value="Overall Accuracy:")
        cls.sheet.cell(row=last_row, column=2, value=f"{accuracy:.2f}%")

        # Save the Excel file
        cls.wb.save("emotion_analysis_results.xlsx")
        print(f"\nSaved results to 'emotion_analysis_results.xlsx'. Overall accuracy: {accuracy:.2f}%")

    def process_scenario_lines(self, scenario_name, lines_with_expected):
        """
        Helper function that processes each line in 'analysis' mode
        for emotion detection and logs the results in the Excel sheet.
        Each item in 'lines_with_expected' is (line_text, expected_emotion).
        """
        for (line_text, expected_emotion) in lines_with_expected:
            # We'll pass this line to the chatbot's emotion detection logic
            # The chatbot usually does stage classification, but we only care
            # about the detected 'dominant_emotion' from the internal process.
            # So let's replicate or re-use logic from 'ChatbotGenerative.process_line'

            # We'll do a simplified approach: detect emotion from 'line_text'
            emotion_results = self.chatbot.emotion_detector.detect_emotion(line_text)
            detected_emotion = emotion_results["label"]

            # Log results in Excel
            row = self.__class__.current_row
            self.__class__.sheet.cell(row=row, column=1, value=scenario_name)
            self.__class__.sheet.cell(row=row, column=2, value=line_text)
            self.__class__.sheet.cell(row=row, column=3, value=expected_emotion)
            self.__class__.sheet.cell(row=row, column=4, value=detected_emotion)

            # Check correctness
            is_correct = (expected_emotion == detected_emotion)
            self.__class__.sheet.cell(row=row, column=5, value="Yes" if is_correct else "No")

            # Update counters
            self.__class__.current_row += 1
            self.__class__.total_lines += 1
            if is_correct:
                self.__class__.correct_count += 1

    def test_forming_scenario_excel(self):
        """
        Example scenario lines for a Forming stage.
        Each tuple is (line_text, expected_emotion).
        """
        forming_lines = [
            ("Alice: Hello everyone! I’m super excited to be part of this team!", "excitement"),
            ("Bob: I’m eager to get started, but a bit anxious about responsibilities.", "anxiety"),
            ("Cara: I’m curious about our overall goals.", "curiosity"),
            ("Dan: I’m enthusiastic, but also uncertain about next steps.", "excitement"),
            # or 'uncertainty' if you prefer
            ("Alice: Just to confirm, I’ll handle the design aspects, right?", "eagerness")
        ]
        self.process_scenario_lines("Forming", forming_lines)

    def test_storming_scenario_excel(self):
        storming_lines = [
            ("Erin: I’m frustrated that we STILL don’t have a final schedule!", "frustration"),
            ("Felix: This tension is getting out of hand.", "tension"),
            ("Gina: I keep feeling defensive...", "defensiveness"),
            ("Henry: I'm uncertain if we can fix this in time.", "uncertainty"),
            ("Felix: I want us to stop fighting and focus on solutions.", "frustration")
        ]
        self.process_scenario_lines("Storming", storming_lines)

    def test_norming_scenario_excel(self):
        norming_lines = [
            ("Iris: I’m relieved we finally assigned tasks.", "relief"),
            ("Jack: I trust we can be more cohesive now.", "trust"),
            ("Kara: This acceptance is refreshing. Camaraderie seems higher than ever!", "acceptance"),
            ("Luke: I’m optimistic about hitting our deadlines.", "optimism")
        ]
        self.process_scenario_lines("Norming", norming_lines)

    def test_performing_scenario_excel(self):
        performing_lines = [
            ("Mia: Great that we overcame earlier issues. I'm confident now!", "confidence"),
            ("Nate: I accept everyone’s feedback. Let’s keep cooperation strong.", "enthusiasm"),  # or "acceptance"
            ("Olivia: I’m proud of how we overcame tensions. Our accomplishments are growing.", "pride"),
            ("Paul: Let’s finalize the milestone. I'm satisfied with our progress!", "satisfaction")
        ]
        self.process_scenario_lines("Performing", performing_lines)

    def test_adjourning_scenario_excel(self):
        adjourning_lines = [
            ("Quinn: Feeling nostalgic knowing we’re wrapping up soon.", "nostalgia"),
            ("Riley: I’m both sad and proud. So much accomplished.", "sadness"),
            ("Sam: Definitely a sense of closure. We should celebrate!", "closure"),
            ("Tina: Let’s document everything we learned before we disband.", "reflection")
        ]
        self.process_scenario_lines("Adjourning", adjourning_lines)
